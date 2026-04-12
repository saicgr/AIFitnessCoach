#!/usr/bin/env python3
"""
Rebuild exercise_library by merging 3 data sources:
1. Local video/illustration folders (exercise_name, video_s3_path, image_s3_path)
2. Excel file (instructions, tips, primary/secondary muscles, equipment)
3. Backup DB export (instructions, equipment, goals, suitable_for, avoid_if, etc.)

Outputs: /tmp/merged_exercises.json with all exercises and fill state
"""

import json
import os
import openpyxl

# === CONFIG ===
VIDEO_DIR = "/Users/saichetangrandhe/Downloads/VERTICAL VIDEOS ALL"
ILLUST_DIR = "/Users/saichetangrandhe/Downloads/ILLUSTRATIONS ALL"
EXCEL_PATH = "/Users/saichetangrandhe/AIFitnessCoach/1500+ exercise data.xlsx"
BACKUP_PATH = "/tmp/backup_exercises.json"
OUTPUT_PATH = "/tmp/merged_exercises.json"

S3_VIDEO_PREFIX = "s3://ai-fitness-coach/VERTICAL VIDEOS ALL"
S3_ILLUST_PREFIX = "s3://ai-fitness-coach/ILLUSTRATIONS ALL"


def load_video_files():
    """Scan local video folder, return list of (exercise_name, folder, filename)."""
    exercises = []
    for folder in sorted(os.listdir(VIDEO_DIR)):
        folder_path = os.path.join(VIDEO_DIR, folder)
        if not os.path.isdir(folder_path):
            continue
        for f in sorted(os.listdir(folder_path)):
            if f.endswith(".mp4"):
                name = f.replace(".mp4", "")
                exercises.append({
                    "exercise_name": name,
                    "folder": folder,
                    "video_filename": f,
                })
    return exercises


def load_illustration_files():
    """Scan local illustration folder, return dict of lowercase_name -> (folder, filename)."""
    illust_map = {}
    for folder in sorted(os.listdir(ILLUST_DIR)):
        folder_path = os.path.join(ILLUST_DIR, folder)
        if not os.path.isdir(folder_path):
            continue
        for f in sorted(os.listdir(folder_path)):
            if f.lower().endswith((".jpg", ".jpeg", ".png")):
                name = f.rsplit(".", 1)[0]
                key = name.lower().strip()
                # Prefer exact match; if duplicate, keep first (already sorted)
                if key not in illust_map:
                    illust_map[key] = (folder, f)
    return illust_map


def load_excel_data():
    """Load Excel file, return dict of lowercase_exercise_name -> data."""
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Sheet1"]

    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [h for h in row if h is not None]

    excel_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        name = data.get("Exercise")
        if name:
            key = name.lower().strip()
            excel_map[key] = {
                "instructions": data.get("Exercise Instructions (step by step)"),
                "tips": data.get("Exercise Tips"),
                "target_muscle": data.get("Primary Activating Muscles"),
                "secondary_muscles_raw": data.get("Secondary Activating Muscles"),
                "equipment": data.get("Equipment"),
            }
    wb.close()
    return excel_map


def load_backup_data():
    """Load backup DB export, return dict of lowercase_exercise_name -> data."""
    with open(BACKUP_PATH, "r") as f:
        backup_list = json.load(f)

    backup_map = {}
    for ex in backup_list:
        key = ex["exercise_name"].lower().strip()
        backup_map[key] = ex
    return backup_map


def parse_secondary_muscles(raw_text):
    """Parse secondary muscles text into a list."""
    if not raw_text:
        return None
    # Split on comma followed by space and uppercase (new muscle group)
    # e.g. "Hamstrings (Biceps Femoris), Shoulders (Deltoids)"
    parts = []
    current = ""
    for char in raw_text:
        current += char
    # Simple split by comma, but keep parenthetical content together
    import re
    # Split on ", " that is NOT inside parentheses
    muscles = re.split(r",\s*(?=[A-Z])", raw_text)
    return [m.strip() for m in muscles if m.strip()]


def merge_all():
    print("Loading video files...")
    videos = load_video_files()
    print(f"  Found {len(videos)} videos")

    print("Loading illustration files...")
    illust_map = load_illustration_files()
    print(f"  Found {len(illust_map)} illustrations")

    print("Loading Excel data...")
    excel_map = load_excel_data()
    print(f"  Found {len(excel_map)} Excel entries")

    print("Loading backup data...")
    backup_map = load_backup_data()
    print(f"  Found {len(backup_map)} backup entries")

    # Build merged exercises
    merged = []
    stats = {
        "total": 0,
        "has_excel": 0,
        "has_backup": 0,
        "has_illustration": 0,
        "has_instructions": 0,
        "has_target_muscle": 0,
        "has_secondary_muscles": 0,
        "has_equipment": 0,
        "has_goals": 0,
        "no_data": 0,
    }

    for vid in videos:
        name = vid["exercise_name"]
        folder = vid["folder"]
        key = name.lower().strip()
        stats["total"] += 1

        # Base row
        exercise = {
            "exercise_name": name,
            "video_s3_path": f"{S3_VIDEO_PREFIX}/{folder}/{vid['video_filename']}",
            "image_s3_path": None,
            "folder": folder,  # metadata for web research agents
            # Columns to populate
            "body_part": None,
            "equipment": None,
            "target_muscle": None,
            "secondary_muscles": None,
            "instructions": None,
            "tips": None,
            "difficulty_level": None,
            "category": None,
            "goals": None,
            "suitable_for": None,
            "avoid_if": None,
            "is_timed": None,
            "is_unilateral": None,
            "single_dumbbell_friendly": None,
            "single_kettlebell_friendly": None,
            "raw_data": None,
            # Source tracking
            "_sources": [],
        }

        # Match illustration
        if key in illust_map:
            ifolder, ifile = illust_map[key]
            exercise["image_s3_path"] = f"{S3_ILLUST_PREFIX}/{ifolder}/{ifile}"
            stats["has_illustration"] += 1

        # Merge Excel data (primary source for muscles, instructions)
        if key in excel_map:
            ex = excel_map[key]
            stats["has_excel"] += 1
            exercise["_sources"].append("excel")

            if ex["instructions"]:
                exercise["instructions"] = ex["instructions"]
                if ex["tips"]:
                    exercise["instructions"] += "\n\nTips:\n" + ex["tips"]
                exercise["tips"] = ex["tips"]
                stats["has_instructions"] += 1

            if ex["target_muscle"]:
                exercise["target_muscle"] = ex["target_muscle"]
                stats["has_target_muscle"] += 1

            if ex["secondary_muscles_raw"]:
                exercise["secondary_muscles"] = parse_secondary_muscles(ex["secondary_muscles_raw"])
                stats["has_secondary_muscles"] += 1

            if ex["equipment"]:
                exercise["equipment"] = ex["equipment"]
                stats["has_equipment"] += 1

        # Gap-fill from backup
        if key in backup_map:
            bk = backup_map[key]
            stats["has_backup"] += 1
            exercise["_sources"].append("backup")

            # Only fill if Excel didn't provide
            if not exercise["instructions"] and bk.get("instructions"):
                exercise["instructions"] = bk["instructions"]
                stats["has_instructions"] += 1

            if not exercise["target_muscle"] and bk.get("target_muscle"):
                exercise["target_muscle"] = bk["target_muscle"]
                stats["has_target_muscle"] += 1

            if not exercise["equipment"] and bk.get("equipment"):
                exercise["equipment"] = bk["equipment"]
                stats["has_equipment"] += 1

            # These only exist in backup
            if bk.get("difficulty_level"):
                exercise["difficulty_level"] = bk["difficulty_level"]
            if bk.get("goals"):
                exercise["goals"] = bk["goals"]
                stats["has_goals"] += 1
            if bk.get("suitable_for"):
                exercise["suitable_for"] = bk["suitable_for"]
            if bk.get("avoid_if"):
                exercise["avoid_if"] = bk["avoid_if"]
            if bk.get("is_timed") is not None:
                exercise["is_timed"] = bk["is_timed"]
            if bk.get("is_unilateral") is not None:
                exercise["is_unilateral"] = bk["is_unilateral"]
            if bk.get("single_dumbbell_friendly") is not None:
                exercise["single_dumbbell_friendly"] = bk["single_dumbbell_friendly"]
            if bk.get("single_kettlebell_friendly") is not None:
                exercise["single_kettlebell_friendly"] = bk["single_kettlebell_friendly"]
            if bk.get("raw_data"):
                exercise["raw_data"] = bk["raw_data"]

        # Track exercises with no data at all
        if not exercise["_sources"]:
            stats["no_data"] += 1

        merged.append(exercise)

    # Save merged data
    with open(OUTPUT_PATH, "w") as f:
        json.dump(merged, f, indent=2, default=str)

    # Print report
    print(f"\n{'='*60}")
    print(f"MERGE REPORT")
    print(f"{'='*60}")
    print(f"Total exercises: {stats['total']}")
    print(f"  With Excel data: {stats['has_excel']} ({stats['has_excel']/stats['total']*100:.0f}%)")
    print(f"  With Backup data: {stats['has_backup']} ({stats['has_backup']/stats['total']*100:.0f}%)")
    print(f"  With Illustration: {stats['has_illustration']} ({stats['has_illustration']/stats['total']*100:.0f}%)")
    print(f"  NO data at all: {stats['no_data']} ({stats['no_data']/stats['total']*100:.0f}%)")
    print(f"\nColumn fill rates after merge:")
    print(f"  instructions: {stats['has_instructions']} ({stats['has_instructions']/stats['total']*100:.0f}%)")
    print(f"  target_muscle: {stats['has_target_muscle']} ({stats['has_target_muscle']/stats['total']*100:.0f}%)")
    print(f"  secondary_muscles: {stats['has_secondary_muscles']} ({stats['has_secondary_muscles']/stats['total']*100:.0f}%)")
    print(f"  equipment: {stats['has_equipment']} ({stats['has_equipment']/stats['total']*100:.0f}%)")
    print(f"  goals: {stats['has_goals']} ({stats['has_goals']/stats['total']*100:.0f}%)")
    print(f"\nStill needs web research:")
    print(f"  category: ALL {stats['total']}")
    print(f"  body_part: ALL {stats['total']}")
    print(f"  difficulty_level, movement_pattern, etc.: most")
    print(f"\nOutput saved to: {OUTPUT_PATH}")

    # Also output the list of exercises with NO data for web research priority
    no_data_exercises = [e for e in merged if not e["_sources"]]
    with open("/tmp/exercises_no_data.json", "w") as f:
        json.dump(no_data_exercises, f, indent=2)
    print(f"Exercises needing full web research: /tmp/exercises_no_data.json ({len(no_data_exercises)} exercises)")

    # Output exercises grouped by folder for agent swarm
    by_folder = {}
    for e in merged:
        folder = e["folder"]
        if folder not in by_folder:
            by_folder[folder] = []
        by_folder[folder].append({
            "exercise_name": e["exercise_name"],
            "target_muscle": e["target_muscle"],
            "secondary_muscles": e["secondary_muscles"],
            "equipment": e["equipment"],
            "has_instructions": e["instructions"] is not None,
        })

    for folder, exercises in by_folder.items():
        path = f"/tmp/exercises_{folder.replace(' ', '_').replace('-', '_')}.json"
        with open(path, "w") as f:
            json.dump(exercises, f, indent=2)
        print(f"  {folder}: {len(exercises)} exercises -> {path}")


if __name__ == "__main__":
    merge_all()
