"""
Shared helpers for creator-program adapters.

Keeping the regexes + openpyxl wrangling in one place buys us:

  * A single source of truth for AMRAP / RPE / percent / rep-range encoding —
    creators use wildly different text ("5+", "AMRAP", "x3+", "@RPE 8",
    "75-80%", "RIR 2"…), but they all map to the same CanonicalProgramTemplate
    shape.
  * One tested openpyxl loader that handles .xlsx AND .xlsm, disables VBA,
    reads computed values (not formulas), and iterates hidden tabs too —
    required for Wendler's hidden `Calc` sheet and nSuns' hidden `LOG` sheet.
  * A single implementation of the "did the user fill this in?" heuristic
    that every adapter can call after parsing to decide whether to also emit
    CanonicalSetRow history alongside the CanonicalProgramTemplate plan.

If you add a new creator adapter, prefer to grow this file rather than copy
its helpers into your module — every adapter benefits from a fix here.
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Iterable, List, Optional, Tuple
from uuid import UUID

from core.logger import get_logger

from ..canonical import (
    CanonicalProgramTemplate,
    CanonicalSetRow,
    LoadPrescription,
    LoadPrescriptionKind,
    ParseResult,
    PrescribedDay,
    PrescribedExercise,
    PrescribedSet,
    PrescribedWeek,
    RepTarget,
    SetType,
    WeightUnit,
    convert_to_kg,
    parse_eu_decimal,
)

logger = get_logger(__name__)


# ─────────────────────────── Regex patterns ───────────────────────────

# Percent-of-1RM / percent-of-TM. Accepts "75%", "75-80%", "75 - 80%", "80 %".
_RE_PERCENT = re.compile(
    r"(\d{1,3}(?:\.\d+)?)\s*(?:-|to|–|—)\s*(\d{1,3}(?:\.\d+)?)\s*%"
    r"|(\d{1,3}(?:\.\d+)?)\s*%",
    re.IGNORECASE,
)

# Rep targets. Accepts "8", "8-12", "5+" (AMRAP), "AMRAP", "Max", "x1+".
_RE_REP_RANGE = re.compile(
    r"^\s*(\d{1,3})\s*(?:-|to|–|—)\s*(\d{1,3})\s*\+?\s*$",
    re.IGNORECASE,
)
_RE_REP_AMRAP_SUFFIX = re.compile(r"^\s*x?\s*(\d{1,3})\s*\+\s*$", re.IGNORECASE)
_RE_REP_SINGLE = re.compile(r"^\s*(\d{1,3})\s*$")
_RE_REP_AMRAP_WORDS = re.compile(
    r"^\s*(amrap|max|to\s*failure|tf|f|fail)\s*$", re.IGNORECASE
)

# RPE / RIR encodings. "@8", "RPE 8", "@7.5", "8-9", "RIR 2", "@2 RIR".
_RE_RPE = re.compile(
    r"(?:@\s*)?(?:rpe\s*)?(\d{1,2}(?:\.\d+)?)"
    r"(?:\s*(?:-|to|–|—)\s*(\d{1,2}(?:\.\d+)?))?",
    re.IGNORECASE,
)
_RE_RIR = re.compile(
    r"(?:@\s*)?(?:rir\s*)?(\d{1,2})"
    r"(?:\s*(?:-|to|–|—)\s*(\d{1,2}))?\s*rir",
    re.IGNORECASE,
)

# Weight-like numeric string (with kg/lb suffix or bare number).
_RE_WEIGHT = re.compile(
    r"([-+]?\d+(?:[.,]\d+)?)\s*(kg|lb|lbs|kilos|kilogramme|pounds?|#)?",
    re.IGNORECASE,
)

# "5×5", "3x10", "3 × 8", "5/5/5/5/3" compound patterns.
_RE_SETS_X_REPS = re.compile(
    r"(\d{1,2})\s*[x×]\s*(\d{1,3})", re.IGNORECASE
)
_RE_SLASH_REPS = re.compile(r"^\s*\d+(?:\s*[/,]\s*\d+)+\s*\+?\s*$")

# Tempo "3-1-1-0", "@3010".
_RE_TEMPO = re.compile(r"^\s*@?\s*(\d{1}(?:\s*-\s*\d{1}){2,3})\s*$")

# Rough date detector for "is this a real calendar date?" signal.
_RE_DATE_LIKE = re.compile(
    r"(\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4})"
    r"|(\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+\d{1,2}\b)",
    re.IGNORECASE,
)

# ─────────────────────────── Primitives ───────────────────────────

AMRAP_TOKENS = {"amrap", "max", "failure", "tf", "f", "fail", "to failure"}


def normalize_unit_hint(raw: Optional[str]) -> WeightUnit:
    """Map a freeform unit string → WeightUnit.

    Recognizes all the typo variants I've seen in the wild ("kg", "KG", "Kilos",
    "lb", "Lbs", "LB", "pounds", "#"). Unknown input defaults to KG — the
    canonical model stores kg internally.
    """
    if raw is None:
        return WeightUnit.KG
    s = str(raw).strip().lower()
    if not s:
        return WeightUnit.KG
    if s in {"kg", "kilo", "kilos", "kilogramme", "kilogram", "kilograms"}:
        return WeightUnit.KG
    if s in {"lb", "lbs", "pound", "pounds", "#"}:
        return WeightUnit.LB
    if s in {"st", "stone", "stones"}:
        return WeightUnit.STONE
    return WeightUnit.KG


def parse_percent(cell: Any) -> Optional[Tuple[float, float]]:
    """Return (min_fraction, max_fraction) in [0,1] from a cell containing
    '75%', '75-80%', or a bare float 0.75. Returns None on blank / non-percent.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        v = float(cell)
        # Accept bare ratios (0..1) AND bare percents (1..100).
        if 0.0 < v <= 1.5:
            return (v, v)
        if 1.5 < v <= 100.0:
            return (v / 100.0, v / 100.0)
        return None
    s = str(cell).strip()
    if not s:
        return None
    m = _RE_PERCENT.search(s)
    if not m:
        return None
    if m.group(1) and m.group(2):
        lo = float(m.group(1)) / 100.0
        hi = float(m.group(2)) / 100.0
        return (lo, hi)
    if m.group(3):
        v = float(m.group(3)) / 100.0
        return (v, v)
    return None


def parse_rep_target(cell: Any) -> Tuple[Optional[RepTarget], bool]:
    """Return (RepTarget, is_amrap_last).

    Handles: 8, 8-12, 5+, x5+, AMRAP, MAX, "to failure", "F".
    `is_amrap_last` means the final set of the prescription is AMRAP —
    the rest of the sets use rep_target.min..max.
    """
    if cell is None:
        return None, False
    if isinstance(cell, (int, float)):
        v = int(cell)
        return RepTarget(min=v, max=v, amrap_last=False), False

    s = str(cell).strip()
    if not s:
        return None, False

    # "AMRAP" / "max" / "F" alone.
    if _RE_REP_AMRAP_WORDS.match(s):
        # Single AMRAP set — use 1 as lower bound (arbitrary) and True flag.
        return RepTarget(min=1, max=99, amrap_last=True), True

    # "5+" / "x5+" — AMRAP suffix, still honors the stated minimum.
    m = _RE_REP_AMRAP_SUFFIX.match(s)
    if m:
        v = int(m.group(1))
        return RepTarget(min=v, max=99, amrap_last=True), True

    # "8-12" / "3-5" / "8-12+".
    m = _RE_REP_RANGE.match(s)
    if m:
        lo = int(m.group(1))
        hi = int(m.group(2))
        amrap = s.rstrip().endswith("+")
        return RepTarget(min=lo, max=hi, amrap_last=amrap), amrap

    # Plain "8".
    m = _RE_REP_SINGLE.match(s)
    if m:
        v = int(m.group(1))
        return RepTarget(min=v, max=v, amrap_last=False), False

    # StrongLifts / Metallicadpa slash encoding. "5/5/5/5/3" — static.
    # "5/5+" — Metallicadpa-style compound: (N-1) fixed + last-set AMRAP.
    if _RE_SLASH_REPS.match(s):
        trailing_plus = s.rstrip().endswith("+")
        parts = [p.strip().rstrip("+") for p in re.split(r"[/,]", s) if p.strip()]
        nums = [int(p) for p in parts if p.isdigit()]
        if nums:
            return (
                RepTarget(
                    min=min(nums),
                    max=(99 if trailing_plus else max(nums)),
                    amrap_last=trailing_plus,
                ),
                trailing_plus,
            )

    # Unrecognized.
    return None, False


def parse_rpe_or_rir(cell: Any) -> Tuple[Optional[float], Optional[int]]:
    """Pull RPE (float) or RIR (int). Returns (rpe, rir) — at most one is
    non-None. Handles ranges by taking the upper bound (more conservative load
    cue) as a single scalar.
    """
    if cell is None:
        return None, None
    if isinstance(cell, (int, float)):
        v = float(cell)
        # Convention: if it's <= 5 and integer-like, treat as RIR; else RPE.
        if v <= 5 and abs(v - round(v)) < 1e-9:
            return None, int(round(v))
        return v, None
    s = str(cell).strip()
    if not s:
        return None, None
    low = s.lower()
    # RIR explicit token.
    m = _RE_RIR.search(low)
    if m:
        hi = m.group(2) or m.group(1)
        return None, int(hi)
    if "rpe" in low or "@" in low:
        m2 = _RE_RPE.search(low)
        if m2:
            hi = m2.group(2) or m2.group(1)
            return float(hi), None
    # Bare number like "8" or "7-8" interpreted as RPE when this column is
    # RPE-labeled; caller invokes with column context.
    m3 = _RE_RPE.search(low)
    if m3:
        hi = m3.group(2) or m3.group(1)
        try:
            return float(hi), None
        except ValueError:
            pass
    return None, None


def safe_int(cell: Any) -> Optional[int]:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        if float(cell).is_integer():
            return int(cell)
        return int(round(float(cell)))
    s = str(cell).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def safe_float(cell: Any) -> Optional[float]:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    return parse_eu_decimal(str(cell))


def parse_weight_cell(cell: Any, unit_default: WeightUnit) -> Optional[float]:
    """Return kg value (float) for a weight cell. Preserves the unit hint when
    the cell itself has no unit suffix. Returns None for blank / non-numeric.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return convert_to_kg(float(cell), unit_default)
    s = str(cell).strip()
    if not s:
        return None
    m = _RE_WEIGHT.search(s)
    if not m:
        return None
    raw_val = parse_eu_decimal(m.group(1))
    if raw_val is None:
        return None
    unit_str = (m.group(2) or "").lower().strip()
    unit = unit_default
    if unit_str in {"kg", "kilos", "kilogramme"}:
        unit = WeightUnit.KG
    elif unit_str in {"lb", "lbs", "pounds", "pound", "#"}:
        unit = WeightUnit.LB
    return convert_to_kg(raw_val, unit)


# ─────────────────────────── Sheet loaders ───────────────────────────

def load_workbook_from_bytes(data: bytes):
    """Open an XLSX/XLSM from raw bytes with sensible defaults for creator
    sheets: read computed values (not formulas), disable VBA, keep hidden
    sheets accessible."""
    # Import lazily so test environments without openpyxl can still import
    # the module for type annotations.
    import openpyxl  # type: ignore

    return openpyxl.load_workbook(
        io.BytesIO(data),
        data_only=True,        # resolve formulas → computed values
        keep_vba=False,        # ignore VBA even in .xlsm
        read_only=False,       # we iterate ranges and named ranges
    )


def iter_all_sheets(workbook) -> Iterable:
    """Iterate every sheet, including hidden ones (.sheet_state == 'hidden').
    openpyxl's default iteration already includes them, but we expose this
    explicitly for clarity — Wendler's Calc tab and nSuns' LOG tab are hidden
    by default and must be followed to resolve formulas."""
    for ws in workbook.worksheets:
        yield ws


def resolve_merged_cell(worksheet, row: int, col: int) -> Any:
    """Return the value of a merged cell given ANY coordinate inside the merge.

    openpyxl only stores the value in the anchor cell (top-left) — all other
    cells in the merge return None. This walks the merge ranges and returns
    the anchor value when (row, col) falls inside one.
    """
    cell = worksheet.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mr in worksheet.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return worksheet.cell(row=mr.min_row, column=mr.min_col).value
    return None


def find_named_range(workbook, name: str) -> Optional[Any]:
    """Resolve a workbook-level named range (e.g. Wendler's `TM_Squat`).
    Returns the raw cell value or None."""
    try:
        defined = workbook.defined_names.get(name)
    except Exception:
        return None
    if defined is None:
        return None
    try:
        dests = list(defined.destinations)
    except Exception:
        return None
    if not dests:
        return None
    sheet_name, coord = dests[0]
    ws = workbook[sheet_name]
    return ws[coord].value


def sheet_text_pool(workbook) -> str:
    """All-lowercase concatenation of every string cell in the workbook.
    Used by the TemplateClassifier copyright-header signal + creator
    fingerprint secondary checks inside adapters. Capped at 1MB to avoid
    blowing up on pathological sheets."""
    chunks: list[str] = []
    total = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                if isinstance(v, str):
                    chunks.append(v.lower())
                    total += len(v)
                    if total > 1_000_000:
                        return " ".join(chunks)
    return " ".join(chunks)


# ─────────────────────────── Template classification signals ───────────────────────────

@dataclass
class SheetSignals:
    """Observations an adapter collects during parse that feed into
    TemplateClassifier.score(...). A single struct keeps the call site tidy."""
    date_fill_ratio: float = 0.0
    weight_fill_ratio: float = 0.0
    formula_density: float = 0.0
    has_prescribed_and_achieved_cols: bool = False
    only_prescribed_filled: bool = False
    static_weight_across_weeks: bool = False
    monotonic_across_weeks: bool = False
    has_protected_cells: bool = False
    has_single_1rm_input: bool = False
    tab_names_are_weeks: bool = False
    tab_names_are_dates: bool = False
    has_copyright_header: bool = False
    notes_are_prescription_style: bool = False
    notes_are_reflection_style: bool = False


def collect_signals_from_workbook(workbook, creator_needles: Iterable[str] = ()) -> SheetSignals:
    """Populate a SheetSignals struct by walking the workbook once.

    We approximate every signal from observable cell contents — we never
    assume a specific schema here so this is safe to call on any spreadsheet.
    """
    tab_names = [ws.title for ws in workbook.worksheets]
    tab_joined_lower = " ".join(tab_names).lower()
    tab_names_are_weeks = any(
        re.search(r"\bweek\s*\d+\b", t, re.IGNORECASE) for t in tab_names
    )
    tab_names_are_dates = any(_RE_DATE_LIKE.search(t) for t in tab_names)

    text_pool = sheet_text_pool(workbook)
    has_copyright_header = any(n.lower() in text_pool for n in creator_needles) or (
        "©" in text_pool or "copyright" in text_pool
    )
    has_single_1rm_input = bool(
        re.search(r"\b1\s*rm\b|\btraining\s*max\b|\btm\b", text_pool)
    )

    # Fill ratio approximations: sample first 200 rows of first sheet.
    first = workbook.worksheets[0]
    date_hits = 0
    weight_hits = 0
    formula_hits = 0
    observed = 0
    for row in first.iter_rows(max_row=200, values_only=False):
        for cell in row:
            observed += 1
            val = cell.value
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            if isinstance(val, datetime):
                date_hits += 1
            elif isinstance(val, str) and _RE_DATE_LIKE.search(val):
                date_hits += 1
            elif isinstance(val, (int, float)):
                # Any numeric cell is a candidate "weight" for this rough signal.
                weight_hits += 1
            if isinstance(val, str) and val.startswith("="):
                # data_only=True normally resolves this away, but keep the
                # check for engines that stored raw formulas.
                formula_hits += 1
    observed_nz = max(observed, 1)

    return SheetSignals(
        date_fill_ratio=date_hits / observed_nz,
        weight_fill_ratio=weight_hits / observed_nz,
        formula_density=formula_hits / observed_nz,
        has_copyright_header=has_copyright_header,
        has_single_1rm_input=has_single_1rm_input,
        tab_names_are_weeks=tab_names_are_weeks,
        tab_names_are_dates=tab_names_are_dates,
    )


# ─────────────────────────── Row-hash helper ───────────────────────────

def deterministic_row_hash(
    *,
    user_id: UUID,
    source_app: str,
    performed_at: datetime,
    exercise_name_canonical: str,
    set_number: Optional[int],
    weight_kg: Optional[float],
    reps: Optional[int],
) -> str:
    """Thin wrapper around CanonicalSetRow.compute_row_hash so adapters stay
    decoupled from that import path."""
    return CanonicalSetRow.compute_row_hash(
        user_id=user_id,
        source_app=source_app,
        performed_at=performed_at,
        exercise_name_canonical=exercise_name_canonical,
        set_number=set_number,
        weight_kg=weight_kg,
        reps=reps,
    )


# ─────────────────────────── Empty / stub result ───────────────────────────

def empty_parse_result(source_app: str, *, warnings: Optional[List[str]] = None) -> ParseResult:
    """Helper for adapters to bail out cleanly when the file is malformed."""
    from ..canonical import ImportMode  # local to keep top-line imports thin
    return ParseResult(
        mode=ImportMode.AMBIGUOUS,
        source_app=source_app,
        warnings=warnings or ["adapter produced no rows"],
    )


def build_template(
    *,
    user_id: UUID,
    source_app: str,
    program_name: str,
    program_creator: Optional[str],
    total_weeks: int,
    days_per_week: int,
    unit_hint: WeightUnit,
    weeks: List[PrescribedWeek],
    one_rm_inputs: Optional[dict] = None,
    training_max_factor: float = 1.0,
    rounding_multiple_kg: float = 2.5,
    program_version: Optional[str] = None,
    body_weight_kg: Optional[float] = None,
    notes: Optional[str] = None,
) -> CanonicalProgramTemplate:
    """Helper that constructs a CanonicalProgramTemplate with the adapter-wide
    defaults already applied. Forwards every kwarg straight through to the
    pydantic model — wrapper exists purely so an adapter can
    `return build_template(...)` in one line.
    """
    return CanonicalProgramTemplate(
        user_id=user_id,
        source_app=source_app,
        program_name=program_name,
        program_creator=program_creator,
        program_version=program_version,
        total_weeks=total_weeks,
        days_per_week=days_per_week,
        unit_hint=unit_hint,
        one_rm_inputs=one_rm_inputs or {},
        body_weight_kg=body_weight_kg,
        rounding_multiple_kg=rounding_multiple_kg,
        training_max_factor=training_max_factor,
        weeks=weeks,
        notes=notes,
    )


def simple_percent_prescription(
    lo: float, hi: float, reference: Optional[str] = None,
    kind: LoadPrescriptionKind = LoadPrescriptionKind.PERCENT_1RM,
) -> LoadPrescription:
    return LoadPrescription(
        kind=kind,
        value_min=lo,
        value_max=hi,
        reference_1rm_exercise=reference,
    )


def absolute_kg_prescription(weight_kg: float) -> LoadPrescription:
    return LoadPrescription(
        kind=LoadPrescriptionKind.ABSOLUTE_KG,
        value_min=weight_kg,
        value_max=weight_kg,
    )


def bodyweight_prescription() -> LoadPrescription:
    return LoadPrescription(kind=LoadPrescriptionKind.BODYWEIGHT)


def rpe_prescription(rpe: float) -> LoadPrescription:
    return LoadPrescription(
        kind=LoadPrescriptionKind.RPE_TARGET,
        value_min=rpe,
        value_max=rpe,
    )


def unspecified_prescription() -> LoadPrescription:
    return LoadPrescription(kind=LoadPrescriptionKind.UNSPECIFIED)


# ─────────────────────────── Filled-history detection ───────────────────────────

def collect_filled_history_rows(
    *,
    user_id: UUID,
    source_app: str,
    performed_at_fallback: datetime,
    template: CanonicalProgramTemplate,
    filled_rows: List[Tuple[str, Optional[float], Optional[int], Optional[int], Optional[str]]],
) -> List[CanonicalSetRow]:
    """Build CanonicalSetRow instances for any sets the user actually logged
    (weight_kg + reps populated). Each input tuple is:

        (exercise_name_raw, weight_kg, reps, set_number, workout_name)

    performed_at_fallback is used when the adapter cannot derive a session
    date from the sheet — calling code is expected to pass the upload time.
    The returned rows are safe to dedup via (user_id, source_row_hash)."""
    rows: List[CanonicalSetRow] = []
    if performed_at_fallback.tzinfo is None:
        performed_at_fallback = performed_at_fallback.replace(tzinfo=timezone.utc)
    for (name, w, r, sn, wn) in filled_rows:
        if w is None and r is None:
            continue  # skip completely-empty entries
        rh = deterministic_row_hash(
            user_id=user_id,
            source_app=source_app,
            performed_at=performed_at_fallback,
            exercise_name_canonical=name.strip().lower(),
            set_number=sn,
            weight_kg=w,
            reps=r,
        )
        rows.append(
            CanonicalSetRow(
                user_id=user_id,
                performed_at=performed_at_fallback,
                workout_name=wn,
                exercise_name_raw=name,
                exercise_name_canonical=None,   # resolver fills this in service.py
                set_number=sn,
                set_type=SetType.WORKING,
                weight_kg=w,
                reps=r,
                source_app=source_app,
                source_row_hash=rh,
            )
        )
    return rows
