"""
Shared fixtures for creator-program adapter tests.

Rather than ship 15+ real creator spreadsheets (most of them paid), we build
minimal .xlsx files in-memory at collection time that capture the essential
shape of each adapter's input. Adapters are tested against these synthetic
fixtures — real-file regressions live behind a `@pytest.mark.real_fixture`
gate that developers run manually when they have access to the original
spreadsheets.

Every builder function here returns bytes you can feed straight into the
adapter's parse() call without touching the filesystem.
"""
from __future__ import annotations

import io
from pathlib import Path
from uuid import UUID

import pytest

try:
    import openpyxl
except ImportError:  # pragma: no cover — openpyxl is a hard dependency
    openpyxl = None


FIXTURE_DIR = Path(__file__).resolve().parents[3] / "fixtures" / "workout_imports" / "programs"
TEST_USER_ID = UUID("11111111-1111-1111-1111-111111111112")


@pytest.fixture
def test_user_id() -> UUID:
    return TEST_USER_ID


@pytest.fixture
def fixture_dir() -> Path:
    return FIXTURE_DIR


def _write_workbook_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────── Creator-specific builders ───────────────────────────

def build_nippard_xlsx() -> bytes:
    """Single-split Nippard sheet: one tab, Week 1 block only, two exercises."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Body 1 - Squat, OHP"
    # Header area with unit + 1RM inputs.
    ws["A1"] = "© Jeff Nippard"
    ws["A2"] = "Unit"
    ws["B2"] = "kg"
    ws["A3"] = "Squat"
    ws["B3"] = 140
    ws["A4"] = "Bench"
    ws["B4"] = 100
    ws["A5"] = "Deadlift"
    ws["B5"] = 180
    ws["A6"] = "OHP"
    ws["B6"] = 60
    # Week banner.
    ws["A8"] = "WEEK 1"
    # Column headers.
    headers = [
        "Exercise", "Warm-up Sets", "Working Sets", "Reps", "Load",
        "%1RM", "RPE", "Rest", "Notes", "Weight Used", "Reps Achieved",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=10, column=i, value=h)
    # Exercise rows.
    ws.append([])  # row 11 blank
    ws["A12"] = "Back Squat"
    ws["B12"] = 4   # warmup sets
    ws["C12"] = 3   # working sets
    ws["D12"] = "5"
    ws["F12"] = "75-80%"
    ws["G12"] = 7.5
    ws["H12"] = "3-4 min"
    ws["I12"] = "Focus on technique."
    ws["J12"] = 105  # weight used — history signal
    ws["K12"] = 5    # reps achieved
    ws["A13"] = "Overhead Press"
    ws["C13"] = 3
    ws["D13"] = "8-10"
    ws["G13"] = 8
    return _write_workbook_to_bytes(wb)


def build_rp_xlsx() -> bytes:
    """RP Male Physique Template — two weeks, one day."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upper 1"
    ws["A1"] = "Renaissance Periodization — Male Physique Template 2.0"
    ws["A2"] = "Week 1"
    headers = [
        "Exercise", "Sets", "Reps", "Load", "RIR",
        "Pump rating", "Performance rating", "Notes",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    ws["A4"] = "Bench Press"
    ws["B4"] = 3
    ws["C4"] = "8-10"
    ws["E4"] = 3
    ws["A5"] = "Barbell Row"
    ws["B5"] = 3
    ws["C5"] = "8-10"
    ws["E5"] = 3
    ws["A7"] = "Week 2"
    for i, h in enumerate(headers, start=1):
        ws.cell(row=8, column=i, value=h)
    ws["A9"] = "Bench Press"
    ws["B9"] = 3
    ws["C9"] = "8-10"
    ws["E9"] = 2
    return _write_workbook_to_bytes(wb)


def build_nuckols_sbs_xlsx() -> bytes:
    """SBS 28 Programs — one lift tab, two weeks."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Squat 3-day Medium"
    ws["A1"] = "Greg Nuckols — Stronger By Science"
    headers = [
        "Week", "Day", "Sets Prescribed", "Reps Prescribed", "%1RM",
        "Sets Completed", "Reps on Last Set",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    ws.append([1, 1, 5, 3, 0.80, None, None])
    ws.append([1, 2, 4, 5, 0.75, None, None])
    ws.append([2, 1, 5, 3, 0.825, 5, 4])   # history signal
    return _write_workbook_to_bytes(wb)


def build_wendler_xlsx() -> bytes:
    """Wendler Poteto-style sheet. We only need TM values to seed the
    template; the program structure is canonical and emitted by the adapter
    without reading cell-by-cell."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws["A1"] = "Jim Wendler 5/3/1 — Poteto v1.28"
    ws["A2"] = "Squat TM"
    ws["B2"] = 126
    ws["A3"] = "Bench TM"
    ws["B3"] = 90
    ws["A4"] = "Deadlift TM"
    ws["B4"] = 162
    ws["A5"] = "OHP TM"
    ws["B5"] = 54
    return _write_workbook_to_bytes(wb)


def build_nsuns_xlsx() -> bytes:
    """nSuns Day 1 (Bench)."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day 1 Bench"
    ws["A1"] = "nSuns 5/3/1"
    ws["A2"] = "Bench 1RM"
    ws["B2"] = 100
    headers = ["Exercise", "Set", "Reps", "%TM", "Weight"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    ws.append(["Bench Press", 1, "x8", 0.65, None])
    ws.append([None, 2, "x6", 0.75, None])
    ws.append([None, 3, "x4", 0.85, None])
    ws.append([None, 4, "x4", 0.85, None])
    ws.append([None, 5, "x5", 0.80, None])
    ws.append([None, 6, "x6", 0.75, None])
    ws.append([None, 7, "x7", 0.70, None])
    ws.append([None, 8, "x1+", 0.65, None])
    return _write_workbook_to_bytes(wb)


def build_gzclp_xlsx() -> bytes:
    """Minimal GZCLP input sheet."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws["A1"] = "GZCLP — Cody LeFever"
    ws["A2"] = "Squat 1RM"
    ws["B2"] = 120
    ws["A3"] = "Bench 1RM"
    ws["B3"] = 90
    ws["A4"] = "Deadlift 1RM"
    ws["B4"] = 140
    ws["A5"] = "Press 1RM"
    ws["B5"] = 55
    return _write_workbook_to_bytes(wb)


def build_metallicadpa_xlsx() -> bytes:
    """Metallicadpa PPL — one tab, two exercises with filled history."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Push 1"
    ws["A1"] = "Metallicadpa PPL"
    headers = ["Exercise", "Sets", "Reps", "Weight (calc)",
               "Reps Done", "Actual Weight"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    ws.append(["Bench Press", 4, "5/5+", 80, 5, 82.5])
    ws.append(["OHP", 3, "8", 50, None, None])
    return _write_workbook_to_bytes(wb)


def build_starting_strength_xlsx() -> bytes:
    """SS community sheet — date + lift columns."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log"
    ws["A1"] = "Starting Strength by Mark Rippetoe"
    ws["A2"] = "Date"
    ws["B2"] = "Squat 3x5"
    ws["C2"] = "Bench 3x5"
    ws["D2"] = "Press 3x5"
    ws["E2"] = "Deadlift 1x5"
    ws["F2"] = "Power Clean 5x3"
    ws.append(["2026-01-01", 70, 60, None, None, None])
    ws.append(["2026-01-03", 72.5, None, 45, None, 50])
    return _write_workbook_to_bytes(wb)


def build_stronglifts_csv() -> bytes:
    """StrongLifts CSV export — minimal."""
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Workout date", "Workout number", "Workout name",
        "Program Name", "Body weight", "Exercise", "Sets & Reps", "Weight",
    ])
    writer.writerow([
        "2026-01-01", 1, "Workout A", "StrongLifts 5x5",
        82.5, "Squat", "5/5/5/5/5", 60,
    ])
    writer.writerow([
        "2026-01-01", 1, "Workout A", "StrongLifts 5x5",
        82.5, "Bench Press", "5/5/5/5/3", 50,
    ])
    return buf.getvalue().encode("utf-8")


def build_lyle_gbr_xlsx() -> bytes:
    """Lyle GBR — wide layout, one day, 3 weeks."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upper A"
    headers = ["Exercise", "Sets", "Reps", "Wk1", "Wk2", "Wk3"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    ws.append(["Bench Press", 3, 8, 70, 72.5, 75])
    ws.append(["Barbell Row", 3, 8, 60, 62.5, 65])
    return _write_workbook_to_bytes(wb)


def build_generic_sheet_xlsx() -> bytes:
    """Generic gym-tracker template."""
    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day 1"
    headers = ["Exercise", "Set 1", "Set 2", "Set 3", "Set 4", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    ws.append(["Bench Press", "8×60kg", "8×65kg", "8×65kg", "6×70kg", ""])
    ws.append(["Lat Pulldown", "10×50kg", "10×55kg", "10×55kg", "", ""])
    return _write_workbook_to_bytes(wb)


def build_minimal_pdf() -> bytes:
    """A tiny valid PDF containing a single line of text. Enough to exercise
    the pypdf path in `_pdf_vision.extract_program_from_pdf`; the Gemini path
    is mocked separately in each test."""
    # Minimal 1-page PDF hand-assembled (saves a dependency on reportlab).
    return (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
        b"/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n"
        b"4 0 obj<</Length 44>>stream\n"
        b"BT /F1 12 Tf 72 720 Td (Sample program PDF) Tj ET\n"
        b"endstream endobj\n"
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n"
        b"xref\n0 6\n0000000000 65535 f\n"
        b"0000000010 00000 n\n0000000054 00000 n\n"
        b"0000000099 00000 n\n0000000180 00000 n\n0000000260 00000 n\n"
        b"trailer<</Size 6/Root 1 0 R>>\n"
        b"startxref\n320\n%%EOF\n"
    )
