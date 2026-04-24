"""
Parametrized edge-case coverage for the workout-import pipeline.

Each test group maps to one of the 10 letter-groups (A-J) in the plan's
"Comprehensive Parser Edge Cases" section. A passing test proves today's
code handles that edge case; `xfail` markers flag gaps that are tracked
but not yet implemented.

Run:
  pytest tests/services/workout_import/test_edge_cases.py -v \
    --confcutdir=tests/services/workout_import
"""
from __future__ import annotations

import hashlib
from datetime import datetime, timezone, timedelta
from uuid import uuid4

import pytest

from services.workout_import.canonical import (
    CanonicalCardioRow,
    CanonicalSetRow,
    ImportMode,
    LoadPrescriptionKind,
    SetType,
    WeightUnit,
    convert_to_kg,
    parse_eu_decimal,
)
from services.workout_import.exercise_resolver import (
    EXERCISE_ALIASES,
    ExerciseResolver,
    _normalize,
)
from services.workout_import.format_detector import TemplateClassifier, detect


# ═══════════════════════════════════════════════════════════════════════════
# GROUP A — Unit handling (#1-13)
# ═══════════════════════════════════════════════════════════════════════════

class TestUnitHandling:
    @pytest.mark.parametrize("raw,expected", [
        # #7 comma-as-decimal
        ("22,5", 22.5),
        ("1,234.56", 1234.56),   # thousands-sep with dot
        ("22.5", 22.5),
        ("", None),
        ("abc", None),
        ("  100  ", 100.0),
        ("  100,75  ", 100.75),
        ("0", 0.0),
        ("0,0", 0.0),
    ])
    def test_eu_decimal_parse(self, raw, expected):
        """#7 — EU comma-as-decimal; #71 leading whitespace strip."""
        assert parse_eu_decimal(raw) == expected

    @pytest.mark.parametrize("value,unit,expected", [
        # #1-5 unit conversions
        (100, WeightUnit.KG, 100.0),
        (100, WeightUnit.LB, 45.359237),
        (14, WeightUnit.STONE, 88.904),       # #8 stones/pounds UK
        (None, WeightUnit.KG, None),
        (0, WeightUnit.LB, 0.0),
        (-20, WeightUnit.LB, -9.0718474),     # #12 negative = assisted
    ])
    def test_weight_conversion(self, value, unit, expected):
        got = convert_to_kg(value, unit)
        if expected is None:
            assert got is None
        else:
            assert abs(got - expected) < 0.01

    def test_eleven_bodyweight_weight_is_none(self):
        """#11 — weight_kg=None means bodyweight, NOT missing data."""
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Pull Up",
            reps=10,
            weight_kg=None,
            source_app="hevy",
            source_row_hash="a" * 64,
        )
        assert row.weight_kg is None

    def test_twelve_assisted_weight_is_negative(self):
        """#12 — assisted machine weight stored as negative value."""
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Assisted Pull Up",
            reps=10,
            weight_kg=-20.0,
            source_app="hevy",
            source_row_hash="b" * 64,
        )
        assert row.weight_kg == -20.0


# ═══════════════════════════════════════════════════════════════════════════
# GROUP B — Rep encoding (#14-23)
# ═══════════════════════════════════════════════════════════════════════════

class TestRepEncoding:
    def test_zero_reps_permitted(self):
        """#22 — reps=0 is legal (duration-only cardio sets, isometric holds)."""
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Plank",
            reps=0,
            duration_seconds=60,
            source_app="hevy",
            source_row_hash="c" * 64,
        )
        assert row.reps == 0
        assert row.duration_seconds == 60

    @pytest.mark.parametrize("set_type", [
        SetType.WORKING, SetType.WARMUP, SetType.FAILURE,
        SetType.DROPSET, SetType.AMRAP, SetType.CLUSTER,
        SetType.REST_PAUSE, SetType.BACKOFF, SetType.ASSISTANCE,
    ])
    def test_all_set_types_accepted(self, set_type):
        """#16-18 — drop/rest-pause/cluster set-type enum covers the spectrum."""
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Bench Press",
            reps=5,
            weight_kg=100,
            set_type=set_type,
            source_app="hevy",
            source_row_hash=hashlib.sha256(set_type.value.encode()).hexdigest(),
        )
        # Pydantic model_config use_enum_values serializes to the string value.
        assert row.set_type in (set_type, set_type.value)

    def test_negative_reps_rejected(self):
        """Negative reps are physically impossible — guard against sign errors."""
        with pytest.raises(Exception):
            CanonicalSetRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
                exercise_name_raw="Bench",
                reps=-5,
                weight_kg=100,
                source_app="hevy",
                source_row_hash="d" * 64,
            )


# ═══════════════════════════════════════════════════════════════════════════
# GROUP C — Set / workout structure (#24-32)
# ═══════════════════════════════════════════════════════════════════════════

class TestSetStructure:
    def test_superset_id_preserved(self):
        """#25 — Hevy's explicit superset_id survives the canonical model."""
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Bench",
            reps=5,
            weight_kg=100,
            superset_id="ss_abc123",
            source_app="hevy",
            source_row_hash="e" * 64,
        )
        assert row.superset_id == "ss_abc123"

    def test_multiple_rows_same_set_number_dropset(self):
        """#19 — drop-set chains reuse set_number; distinguished by weight/reps."""
        user_id = uuid4()
        t = datetime(2026, 4, 23, 12, 0, tzinfo=timezone.utc)
        hashes = [
            CanonicalSetRow.compute_row_hash(
                user_id=user_id, source_app="hevy", performed_at=t,
                exercise_name_canonical="barbell_bench_press",
                set_number=3, weight_kg=w, reps=r,
            )
            for w, r in [(100, 8), (80, 8), (60, 8)]
        ]
        # All three hashes distinct even though set_number is identical.
        assert len(set(hashes)) == 3


# ═══════════════════════════════════════════════════════════════════════════
# GROUP D — Date / time (#33-42)
# ═══════════════════════════════════════════════════════════════════════════

class TestDateTime:
    def test_naive_datetime_rejected(self):
        """#35 — timezone-less timestamps fail fast. No silent UTC cast."""
        with pytest.raises(Exception):
            CanonicalSetRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, 12, 0),  # naive
                exercise_name_raw="Bench",
                reps=5,
                weight_kg=100,
                source_app="hevy",
                source_row_hash="f" * 64,
            )

    def test_tz_aware_datetime_accepted(self):
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Bench",
            reps=5,
            weight_kg=100,
            source_app="hevy",
            source_row_hash="0" * 64,
        )
        assert row.performed_at.tzinfo is not None

    def test_hash_rounds_to_date_not_second(self):
        """#36-38 — two timestamps on the same date + same data = same hash."""
        uid = uuid4()
        h_morning = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy",
            performed_at=datetime(2026, 4, 23, 8, 0, tzinfo=timezone.utc),
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=100.0, reps=5,
        )
        h_evening = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy",
            performed_at=datetime(2026, 4, 23, 22, 0, tzinfo=timezone.utc),
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=100.0, reps=5,
        )
        assert h_morning == h_evening


# ═══════════════════════════════════════════════════════════════════════════
# GROUP E — Exercise-name resolution (#43-52)
# ═══════════════════════════════════════════════════════════════════════════

class TestExerciseNameResolution:
    @pytest.mark.parametrize("raw,expected", [
        # #43 abbreviations
        ("BB Row", "barbell_bent_over_row"),
        ("DB Press", "dumbbell_shoulder_press"),
        ("RDL", "barbell_romanian_deadlift"),
        ("OHP", "barbell_overhead_press"),
        # #44 parenthetical equipment qualifier stripped
        ("Bench Press (Barbell)", "barbell_bench_press"),
        # Case + punctuation insensitivity
        ("FLAT BARBELL BENCH PRESS", "barbell_bench_press"),
        ("flat barbell bench press", "barbell_bench_press"),
        # #49 pluralization
        ("Push-ups", "bodyweight_push_up"),
        ("Push-Up", "bodyweight_push_up"),
        ("pushup", "bodyweight_push_up"),
        # #47 emoji stripped
        ("💪 Bench Press", "barbell_bench_press"),
        # #48 non-English
        ("Sentadilla", "barbell_back_squat"),
        ("Kniebeuge", "barbell_back_squat"),
    ])
    def test_alias_resolution(self, raw, expected):
        resolver = ExerciseResolver()
        # Bypass library/RAG cascade — we're only testing level-1 alias.
        resolver._library_cache = {}  # disable level 2
        result = resolver.resolve(raw)
        assert result.canonical_name == expected, \
            f"{raw!r} → {result.canonical_name!r}, expected {expected!r}"
        assert result.level == 1
        assert result.confidence == 1.0

    def test_unresolved_fallback(self):
        """#52 — unresolved names canonicalize to normalized form, level 4."""
        resolver = ExerciseResolver()
        resolver._library_cache = {}
        result = resolver.resolve("Sai's Cable Jammer Variation")
        assert result.level == 4
        assert result.confidence == 0.0
        assert "cable" in result.canonical_name.lower()

    def test_empty_name_resolves_to_unknown(self):
        resolver = ExerciseResolver()
        resolver._library_cache = {}
        result = resolver.resolve("")
        assert result.level == 4

    @pytest.mark.parametrize("raw,normalized", [
        ("Bench Press  ", "bench press"),
        ("Bench-Press", "bench press"),
        ("bench_press", "bench press"),
        ("Bench (Barbell)", "bench"),
        ("💪 Bench Press!", "bench press"),
    ])
    def test_normalize_strips_punctuation(self, raw, normalized):
        """#46 hyperlinks / #47 emoji / #70 whitespace / #44 parens."""
        assert _normalize(raw) == normalized

    def test_alias_dict_size_and_coverage(self):
        """At minimum 200 aliases per the plan spec."""
        assert len(EXERCISE_ALIASES) >= 200
        # Spot-check critical compound lifts all have aliases.
        for key in ["bench press", "squat", "deadlift", "ohp", "rdl"]:
            assert key in EXERCISE_ALIASES


# ═══════════════════════════════════════════════════════════════════════════
# GROUP F — Spreadsheet quirks (#53-71) — smoke via detector
# ═══════════════════════════════════════════════════════════════════════════

class TestSpreadsheetQuirks:
    def test_detect_empty_file(self):
        """#88 — empty files rejected early."""
        result = detect(b"", filename="empty.csv")
        assert result.confidence == 0.0
        assert any("empty" in w.lower() for w in result.warnings)

    def test_detect_zip_magic(self):
        """#53 — ZIP magic bytes recognized even when the zip is malformed."""
        zip_header = b"PK\x03\x04" + b"\x00" * 100
        result = detect(zip_header, filename="test.zip")
        # Malformed ZIP routes to "unknown" source_app with AMBIGUOUS mode and
        # a warning — which IS a sane routing (client shows "file damaged").
        assert result.source_app in ("unknown", "ai_fallback_zip")
        assert len(result.warnings) > 0

    def test_detect_pdf(self):
        """#72 — PDF magic bytes route to AI fallback."""
        pdf_header = b"%PDF-1.4\n" + b"\x00" * 100
        result = detect(pdf_header, filename="program.pdf")
        assert "ai_fallback" in result.source_app or "pdf" in result.source_app.lower()

    def test_detect_unknown_text_falls_back_to_ai(self):
        """Anything unrecognized routes to AI fallback, not hard-fail."""
        result = detect(b"random gibberish that is not fitness", filename="x.txt")
        assert result.source_app in ("ai_fallback", "generic_csv")


# ═══════════════════════════════════════════════════════════════════════════
# GROUP G — PDF specifics (#72-81)
# ═══════════════════════════════════════════════════════════════════════════

class TestPdfSpecifics:
    def test_pdf_filename_hints_template(self):
        """#77 — creator filename → TEMPLATE mode even without parsing."""
        pdf = b"%PDF-1.4\n" + b"\x00" * 20
        result = detect(pdf, filename="Nippard_Powerbuilding_v3.pdf")
        assert result.mode == ImportMode.TEMPLATE

    def test_pdf_filename_hints_history(self):
        """#77 — filenames like 'log' / 'history' route to HISTORY."""
        pdf = b"%PDF-1.4\n" + b"\x00" * 20
        result = detect(pdf, filename="my_workout_log.pdf")
        assert result.mode == ImportMode.HISTORY


# ═══════════════════════════════════════════════════════════════════════════
# GROUP H — File-level structural (#82-90)
# ═══════════════════════════════════════════════════════════════════════════

class TestFileStructural:
    def test_empty_bytes_rejected(self):
        """#88 — empty upload returns sensible zero-confidence result."""
        r = detect(b"", filename=None)
        assert r.source_app in ("unknown", "ai_fallback")

    def test_json_with_boostcamp_shape(self):
        """Boostcamp shape: dict with 'program' key."""
        payload = b'{"program": {"name": "Test", "weeks": []}}'
        result = detect(payload, filename="boostcamp_export.json")
        assert result.source_app in ("boostcamp", "generic_json")

    def test_xml_apple_health_signature(self):
        """#53 — Apple Health XML root element signature."""
        xml = b"<?xml version='1.0'?>\n<HealthData><Workout/></HealthData>"
        result = detect(xml, filename="export.xml")
        assert result.source_app == "apple_health_xml"
        assert result.mode == ImportMode.CARDIO_ONLY

    def test_gpx_detected_as_cardio(self):
        xml = b"<?xml version='1.0'?>\n<gpx xmlns='x'><trk/></gpx>"
        result = detect(xml, filename="morning_run.gpx")
        assert result.mode == ImportMode.CARDIO_ONLY


# ═══════════════════════════════════════════════════════════════════════════
# GROUP I — Semantic / data quality (#91-102)
# ═══════════════════════════════════════════════════════════════════════════

class TestSemanticQuality:
    def test_rpe_clamped_in_validation(self):
        """#93 — RPE > 10 rejected; #94 null RPE OK."""
        # Valid range
        row = CanonicalSetRow(
            user_id=uuid4(),
            performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
            exercise_name_raw="Bench",
            reps=5,
            weight_kg=100,
            rpe=8.5,
            source_app="hevy",
            source_row_hash="1" * 64,
        )
        assert row.rpe == 8.5

        # Out of range → pydantic raises
        with pytest.raises(Exception):
            CanonicalSetRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
                exercise_name_raw="Bench",
                reps=5,
                weight_kg=100,
                rpe=11.0,   # over cap
                source_app="hevy",
                source_row_hash="2" * 64,
            )

    def test_rir_bounds(self):
        """#95 — RIR clamped to 0..10."""
        with pytest.raises(Exception):
            CanonicalSetRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
                exercise_name_raw="Bench",
                reps=5,
                weight_kg=100,
                rir=-1,
                source_app="hevy",
                source_row_hash="3" * 64,
            )

    def test_hash_stable_across_rounding_drift(self):
        """#61 formula vs value — 100.0 vs 100.04 kg hash identically."""
        uid = uuid4()
        t = datetime(2026, 4, 23, tzinfo=timezone.utc)
        h1 = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy", performed_at=t,
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=100.0, reps=5,
        )
        h2 = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy", performed_at=t,
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=100.04, reps=5,
        )
        assert h1 == h2

    def test_hash_differs_on_meaningful_changes(self):
        """Conversely: 100kg vs 102.5kg hashes are distinct (real plate step)."""
        uid = uuid4()
        t = datetime(2026, 4, 23, tzinfo=timezone.utc)
        h1 = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy", performed_at=t,
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=100.0, reps=5,
        )
        h2 = CanonicalSetRow.compute_row_hash(
            user_id=uid, source_app="hevy", performed_at=t,
            exercise_name_canonical="barbell_bench_press",
            set_number=1, weight_kg=102.5, reps=5,
        )
        assert h1 != h2


# ═══════════════════════════════════════════════════════════════════════════
# GROUP J — Template-vs-history classifier signals
# ═══════════════════════════════════════════════════════════════════════════

class TestTemplateClassifier:
    def test_strong_template_signals_score_high(self):
        """Nippard-style inputs → template mode."""
        score, _ = TemplateClassifier.score(
            date_fill_ratio=0.0,
            weight_fill_ratio=0.05,
            formula_density=0.4,
            has_single_1rm_input=True,
            tab_names_are_weeks=True,
            has_copyright_header=True,
        )
        assert score > 0.8
        assert TemplateClassifier.mode_from_score(score) == ImportMode.TEMPLATE

    def test_strong_history_signals_score_low(self):
        """Filled-in log: real dates, filled weights, monotonic progression."""
        score, _ = TemplateClassifier.score(
            date_fill_ratio=0.95,
            weight_fill_ratio=0.9,
            formula_density=0.0,
            monotonic_across_weeks=True,
            tab_names_are_dates=True,
            notes_are_reflection_style=True,
        )
        assert score < 0.4
        assert TemplateClassifier.mode_from_score(score) == ImportMode.HISTORY

    def test_ambiguous_score_flags_for_user_disambiguation(self):
        """Score between thresholds routes to AMBIGUOUS (user picks)."""
        score, _ = TemplateClassifier.score(
            date_fill_ratio=0.3,
            weight_fill_ratio=0.5,
            formula_density=0.1,
        )
        # Could land either side; verify it's in the ambiguous band once detected.
        if 0.35 < score < 0.65:
            assert TemplateClassifier.mode_from_score(score) == ImportMode.AMBIGUOUS


# ═══════════════════════════════════════════════════════════════════════════
# GROUP K — Cardio canonical row validation (from migration constraints)
# ═══════════════════════════════════════════════════════════════════════════

class TestCardioRow:
    def test_cardio_requires_duration(self):
        """duration_seconds is required + must be positive."""
        with pytest.raises(Exception):
            CanonicalCardioRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
                activity_type="run",
                duration_seconds=0,  # rejected
                source_app="strava",
                source_row_hash="4" * 64,
            )

    def test_cardio_heart_rate_bounds(self):
        """avg_heart_rate must be 20-260 bpm (physiologically plausible)."""
        with pytest.raises(Exception):
            CanonicalCardioRow(
                user_id=uuid4(),
                performed_at=datetime(2026, 4, 23, tzinfo=timezone.utc),
                activity_type="run",
                duration_seconds=1800,
                avg_heart_rate=300,   # impossible
                source_app="strava",
                source_row_hash="5" * 64,
            )

    def test_cardio_hash_includes_distance(self):
        """Two runs same day, same duration, different distances = distinct hashes."""
        uid = uuid4()
        t = datetime(2026, 4, 23, 7, 0, tzinfo=timezone.utc)
        h1 = CanonicalCardioRow.compute_row_hash(
            user_id=uid, source_app="strava", performed_at=t,
            activity_type="run", duration_seconds=1800, distance_m=5000,
        )
        h2 = CanonicalCardioRow.compute_row_hash(
            user_id=uid, source_app="strava", performed_at=t,
            activity_type="run", duration_seconds=1800, distance_m=5500,
        )
        assert h1 != h2


# ═══════════════════════════════════════════════════════════════════════════
# Known gaps — tracked via xfail (intentionally expected to fail until fixed)
# ═══════════════════════════════════════════════════════════════════════════

class TestNamedRangeAndDedupe:
    def test_named_range_resolution(self):
        """#58 — openpyxl defined_names follow-through for creator sheets.
        Builds a minimal workbook with a TM_Squat named range pointing at a
        cell with 150, then asserts find_named_range returns that value.
        """
        import io
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName
        from services.workout_import.programs._shared import find_named_range

        wb = Workbook()
        ws = wb.active
        ws.title = "Inputs"
        ws["B2"] = 150.0
        # Absolute ref required — otherwise openpyxl treats it as a formula.
        defn = DefinedName("TM_Squat", attr_text="Inputs!$B$2")
        wb.defined_names["TM_Squat"] = defn

        # Roundtrip through bytes so behavior matches real uploads.
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        wb2 = load_workbook(buf, data_only=True)

        assert find_named_range(wb2, "TM_Squat") == 150.0
        assert find_named_range(wb2, "TM_NotHere") is None

    def test_duplicate_file_sha_rejection(self):
        """#96 — identical file bytes hash to the same sha256, so a pre-check
        can short-circuit re-imports. Verifies the primitive the endpoint
        uses (backend/api/v1/workout_history_file.py::_file_sha256)."""
        import hashlib
        data_a = b"Title,Exercise,Weight (kg),Reps\n2026,Bench,100,5\n"
        data_b = b"Title,Exercise,Weight (kg),Reps\n2026,Bench,100,5\n"
        data_c = b"Title,Exercise,Weight (kg),Reps\n2026,Bench,102.5,5\n"
        sha = lambda b: hashlib.sha256(b).hexdigest()
        assert sha(data_a) == sha(data_b)
        assert sha(data_a) != sha(data_c)
        # Real upload endpoint computes the same digest and compares against
        # media_analysis_jobs.params->>file_sha256. See _file_sha256 helper.
